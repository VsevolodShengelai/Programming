from flask import Flask, request
import json
import datetime
import openpyxl
import os.path

app = Flask(__name__)


def write_xls():
    global max_line, buff

    book = openpyxl.load_workbook('data.xlsx')
    sheet = book.active

    for raw in buff:
        for i in range(5):
            sheet.cell(max_line, i + 1).value = raw[i]
        max_line += 1

    book.save('data.xlsx')
    book.close

    buff.clear()


@app.route('/', methods=['POST'])
def index():
    global buff, max_id
    req_json = request.get_json()
    req_time = datetime.datetime.now().time()

    for it in req_json['check']:
        raw = [max_id, req_json['user_id'], req_time, it['item'], it['price']]
        max_id += 1
        print(raw)
        buff.append(raw)

    if len(buff) > 0:
        write_xls()

    return '0'


if __name__ == "__main__":
    global buff, max_id, max_line
    buff = []
    max_id = 1
    max_line = 2


if not os.path.exists('data.xlsx'):
    book = openpyxl.Workbook()
    sheet = book.active

    sheet['A1'].value = 'N'
    sheet['B1'].value = 'User ID'
    sheet['C1'].value = 'Datetime'
    sheet['D1'].value = 'Item'
    sheet['E1'].value = 'Prise'

    book.save('data.xlsx')
    book.close


app.run()
